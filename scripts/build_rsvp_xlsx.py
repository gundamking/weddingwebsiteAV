"""Build the RSVP backend workbook for Google Drive.

Reads rsvp/parties.csv -> rsvp/AV Wedding RSVP.xlsx with two tabs:
  - Parties:    Party ID | Party Name | Guest Count   (from parties.csv, numeric)
  - Responses:  Timestamp | Party ID | Coming Count | Genres | Note  (headers only)

Upload the .xlsx to Google Drive and open as a Google Sheet; the Apps Script
backend writes guest replies into the Responses tab.

Run from repo root:  python scripts/build_rsvp_xlsx.py
Requires: openpyxl.
"""
import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CSV_PATH = 'rsvp/parties.csv'
XLSX_PATH = 'rsvp/AV Wedding RSVP.xlsx'

HEADER_FILL = PatternFill('solid', fgColor='6B4E7A')   # site purple
HEADER_FONT = Font(bold=True, color='FFFFFF')


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build():
    wb = Workbook()

    # --- Parties tab ---
    parties = wb.active
    parties.title = 'Parties'
    parties.append(['Party ID', 'Party Name', 'Guest Count'])
    with open(CSV_PATH, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        rows = 0
        for pid, name, count in reader:
            parties.append([pid, name, int(count)])
            rows += 1
    style_header(parties, 3)
    set_widths(parties, [22, 32, 13])

    # --- Responses tab (empty, headers only) ---
    responses = wb.create_sheet('Responses')
    responses.append(['Timestamp', 'Party ID', 'Coming Count', 'Genres', 'Note'])
    style_header(responses, 5)
    set_widths(responses, [22, 22, 14, 36, 40])

    wb.save(XLSX_PATH)
    print('wrote %s  (%d parties, %.1f KB)' %
          (XLSX_PATH, rows, os.path.getsize(XLSX_PATH) / 1024))


if __name__ == '__main__':
    build()
