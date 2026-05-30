"""Consolidate the private Attendee List.xlsx into a public Parties CSV.

Usage:
    python scripts/build_parties_csv.py "Attendee List.xlsx" rsvp/parties.csv \
        --sheets "Bride Side" "Groom Side"

Output columns: Party ID, Party Name, Guest Count
Only Day 1+2 parties. Choose the correct groom tab via --sheets.
"""
import argparse
import csv
import re
import sys

GUEST_COUNT_COL = 4   # column E (0-indexed) = "Guest Count"
NAME_COL = 0          # column A = "Name"

DROP_NAMES = {'name', 'total guest count', 'total room count'}


def slugify(name):
    s = name.strip().lower()
    s = re.sub(r'[&/]', ' ', s)
    s = re.sub(r'[^a-z0-9]+', '-', s)
    return re.sub(r'-+', '-', s).strip('-')


def consolidate(rows, drop_names=DROP_NAMES):
    """rows: list of {'name': str, 'count': number}. Returns list of
    {'id','name','count'} filtered + deduped (max count), with unique slug ids."""
    best = {}
    order = []
    for r in rows:
        name = (r.get('name') or '').strip()
        if not name or name.lower() in drop_names:
            continue
        try:
            count = int(float(r.get('count') or 0))
        except (TypeError, ValueError):
            continue
        if count <= 0:
            continue
        key = re.sub(r'\s+', ' ', name.lower())
        if key not in best:
            best[key] = {'name': name, 'count': count}
            order.append(key)
        else:
            best[key]['count'] = max(best[key]['count'], count)
    out, used = [], set()
    for key in order:
        rec = best[key]
        base = slugify(rec['name']) or 'party'
        pid, n = base, 2
        while pid in used:
            pid, n = '%s-%d' % (base, n), n + 1
        used.add(pid)
        out.append({'id': pid, 'name': rec['name'], 'count': rec['count']})
    return out


def read_sheet(ws):
    rows = []
    for r in ws.iter_rows(values_only=True):
        name = r[NAME_COL] if len(r) > NAME_COL else None
        count = r[GUEST_COUNT_COL] if len(r) > GUEST_COUNT_COL else None
        rows.append({'name': name, 'count': count})
    return rows


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx')
    ap.add_argument('out_csv')
    ap.add_argument('--sheets', nargs='+', default=['Bride Side', 'Groom Side'])
    args = ap.parse_args(argv)

    import openpyxl
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    rows = []
    for sheet in args.sheets:
        if sheet not in wb.sheetnames:
            print('WARNING: sheet not found: %r' % sheet, file=sys.stderr)
            continue
        rows.extend(read_sheet(wb[sheet]))

    parties = consolidate(rows)
    with open(args.out_csv, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['Party ID', 'Party Name', 'Guest Count'])
        for p in parties:
            w.writerow([p['id'], p['name'], p['count']])
    print('Wrote %d parties to %s' % (len(parties), args.out_csv))


if __name__ == '__main__':
    main()
